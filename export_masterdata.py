import json, sys, os
import openpyxl

SKIP_SHEETS = ['enum_definitions']

def excel_to_json(xlsx_path, output_dir, target_sheets=None):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        if target_sheets and sheet_name not in target_sheets:
            continue

        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        data = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            obj = {}
            for header, value in zip(headers, row):
                if value is None:
                    obj[header] = ''
                elif isinstance(value, float) and value == int(value):
                    obj[header] = int(value)
                else:
                    obj[header] = value
            data.append(obj)

        output_path = f'{output_dir}/{sheet_name}.json'
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f'{sheet_name}: {len(data)} rows -> {output_path}')

if __name__ == '__main__':
    xlsx = sys.argv[1]
    outdir = sys.argv[2]
    os.makedirs(outdir, exist_ok=True)
    sheets = sys.argv[3].split(',') if len(sys.argv) > 3 else None
    excel_to_json(xlsx, outdir, sheets)
